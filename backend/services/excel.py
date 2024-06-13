from openpyxl import load_workbook
from openpyxl.styles import Alignment
from flask import send_file, jsonify
import os

def create_invoice(client, date, products, order_id):
    try:
        template_excel = "src/excel/plantilla.xlsx"
        save_path = f'src/excel/facturas/{str(order_id)}_{date}.xlsx'

        if not os.path.exists("src/excel/facturas"):
            os.makedirs("src/excel/facturas")

        wb = load_workbook(template_excel)
        ws = wb.active

        # Write the client data
        ws['E6'] = client['name']
        ws['E7'] = client['cif']
        ws['E8'] = client['email']
        ws['E9'] = client['address']
        ws['E10'] = client['type']
        ws['D12'] = date

        index = 14
        total_discount = 0
        total_price = 0
        total_to_pay = 0
        for product in products:
            ws['A' + str(index)].alignment = Alignment(horizontal='left', vertical='center')
            ws['A' + str(index)] = product['name']
            ws['C' + str(index)].alignment = Alignment(horizontal='right', vertical='center')
            ws['C' + str(index)] = product['packaging']
            ws['D' + str(index)].alignment = Alignment(horizontal='right', vertical='center')
            ws['D' + str(index)] = product['quantity']
            ws['E' + str(index)].alignment = Alignment(horizontal='right', vertical='center')
            ws['E' + str(index)] = str(round(product['price'], 2)) + "€"
            ws['F' + str(index)].alignment = Alignment(horizontal='right', vertical='center')
            ws['F' + str(index)] = product['discount']
            ws['G' + str(index)].alignment = Alignment(horizontal='right', vertical='center')
            discount = ((product['quantity'] * product['price']) * (100 - product['discount'])) / 100
            total_discount += (product['quantity'] * product['price']) - discount
            total_price += (product['quantity'] * product['price'])
            total_to_pay += discount
            ws['G' + str(index)] = str(round(discount, 2)) + "€"
            index+=1

        # Add the final info on the footer
        ws['G43'] = str(round(total_price, 2)) + "€"
        ws['G44'] = str(round(total_discount, 2)) + "€"
        # The IVA for these types of products is 10% of total price(After the discounts)
        iva = round(total_price, 2) * 0.1
        ws['G45'] = str(round(iva, 2)) + "€"
        ws['G46'] = str(round((total_to_pay + iva),2)) + "€"
        wb.save(save_path)
        return round((total_to_pay + iva),2)

    except Exception:
        return None
    
# Return the invoice file .xlsx if exists or 400
def get_invoice(order_id, date):
    filename = f"{order_id}_{date}.xlsx"
    file_path = os.path.join("src/excel/facturas", filename)

    if not os.path.isfile(file_path):
        return jsonify({"error": "File not found"}), 400

    return send_file(
        file_path,
        as_attachment=True,
        download_name=f'Factura_{date}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )

create_invoice("", "", "", "")